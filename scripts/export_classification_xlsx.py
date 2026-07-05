"""
Part 2, Step 4c: export the required table as an XLSX spreadsheet.

Columns (per project description, p.28):
    repository_id, project_type, project_title, primary_class,
    secondary_class // if any
    no_project_files // No files in project in total

Run from the repo root:
    python3 scripts/export_classification_xlsx.py [output_path]

Default output: data/sq26-classification-table.xlsx
"""

import sys
import sqlite3
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from config import DB_PATH

DEFAULT_OUTPUT = "data/sq26-classification-table.xlsx"

COLUMNS = [
    "repository_id",
    "project_type",
    "project_title",
    "primary_class",
    "secondary_class",
    "no_project_files",
]


def fetch_table(con) -> pd.DataFrame:
    return pd.read_sql_query("""
        SELECT
            p.repository_id   AS repository_id,
            p.type            AS project_type,
            p.title           AS project_title,
            p.primary_class   AS primary_class,
            p.secondary_class AS secondary_class,
            (SELECT COUNT(*) FROM files f WHERE f.project_id = p.id) AS no_project_files
        FROM projects p
        ORDER BY p.repository_id, p.type, p.title
    """, con)


def style_workbook(output_path: str, df: pd.DataFrame):
    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    ws = wb["classification"]

    header_font = Font(bold=True, color="FFFFFF", name="Calibri")
    header_fill = PatternFill("solid", fgColor="4472C4")
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # sensible column widths based on content
    widths = {
        "repository_id": 14,
        "project_type": 16,
        "project_title": 60,
        "primary_class": 14,
        "secondary_class": 16,
        "no_project_files": 16,
    }
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths[col_name]

    # body font + wrap long titles
    body_font = Font(name="Calibri", size=11)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(COLUMNS)):
        for cell in row:
            cell.font = body_font
        row[2].alignment = Alignment(wrap_text=False, vertical="center")  # project_title column

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)


def main():
    output_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_OUTPUT

    con = sqlite3.connect(DB_PATH)
    df = fetch_table(con)
    con.close()

    # keep column order and names exact, per the brief
    df = df[COLUMNS]

    df.to_excel(output_path, index=False, sheet_name="classification")
    style_workbook(output_path, df)

    print(f"Wrote {len(df)} rows to {output_path}")
    print()
    print("project_type breakdown:")
    print(df["project_type"].value_counts().to_string())
    print()
    print(f"rows with no secondary_class: {df['secondary_class'].isna().sum()}")
    print(f"rows with no primary_class:   {df['primary_class'].isna().sum()}")


if __name__ == "__main__":
    main()
